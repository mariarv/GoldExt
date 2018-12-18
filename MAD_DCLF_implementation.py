import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook  # not zero-based indexing!!!!!!!!!!!!!!!!
import time
import os

# get the index of the first element in the array which is bigger than a given value
def getElementIndex(value, array):
    if(value < array[0]):
        return '-INF'
    if(value > array[-1]):
        return '+INF'
    else:
        if(value == array[-1]):
            return len(array)
        for i in range(0, len(array) - 1):
            if(value >= array[i] and value < array[i + 1]):
                return i + 2


fileToOpen = 'ml_96_random6_132_600dLoc_60dCluster_EZ_Ripley.xlsx'
wb = load_workbook(fileToOpen, data_only=True)
#print wb.get_sheet_names()

ws = wb['Ripley raw data']

dist = []
L_est = []
L_est_matrix = []

minDist = 13 # cell number, corresponds to 11 nm
maxDist = 82 # cell number, corresponds to 80 nm

for row in ws.iter_rows(min_row = minDist, max_row = maxDist, min_col = 1, max_col = 1):
    for cell in row:
        dist.append(cell.value)

for row in ws.iter_rows(min_row = minDist, max_row = maxDist, min_col = 2, max_col = 2):
    for cell in row:
        L_est.append(cell.value)

L_est_matrix.append(np.asarray(L_est))

#avg_model = np.average(L_est)
int_model = np.trapz(L_est, dist) # trapezoidal integration

#avg_random = []
int_random = []
numberOfRandoms = 200

for i in range(0, numberOfRandoms):
    tmpRandom = []
    for row in ws.iter_rows(min_row = minDist, max_row = maxDist, min_col = (i+3), max_col = (i+3)):
        for cell in row:
            tmpRandom.append(cell.value)

    L_est_matrix.append(np.asarray(tmpRandom))
    #avg_random.append(np.average(tmpRandom))
    int_random.append(np.trapz(tmpRandom, dist))

L_theo_est = np.average(np.transpose(np.asarray(L_est_matrix)),1)

# MAD & DCLF statistics
MAD_data = np.amax(abs(L_est - L_theo_est))
DCLF_data = np.sum((L_est - L_theo_est)**2)

MAD_random = []
DCLF_random = []
for i in range(0, numberOfRandoms):
    MAD_random.append(np.amax(abs(L_est_matrix[i+1] - L_theo_est)))
    DCLF_random.append(np.sum((L_est_matrix[i+1] - L_theo_est)**2))

#index = getElementIndex(avg_model, np.sort(avg_random))
index_MAD = getElementIndex(MAD_data, np.sort(MAD_random))
index_DCLF = getElementIndex(DCLF_data, np.sort(DCLF_random))

print index_MAD, index_DCLF

